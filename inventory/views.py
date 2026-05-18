from rest_framework import viewsets, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import FileResponse
from .models import (
    Asset, Employee, Department, AssetHistory, AssetAssignment, InspectionLog,
    AssetActionRequest, HealthCheckSession, HealthCheckResponse
)
from .serializers import (
    AssetSerializer, EmployeeSerializer, DepartmentSerializer, 
    AssetAssignmentSerializer, InspectionLogSerializer, AssetListSerializer, 
    AssetDetailSerializer, UserSerializer, AssetActionRequestSerializer,
    HealthCheckSessionSerializer, HealthCheckResponseSerializer
)
from rest_framework.decorators import action
from django.db import transaction
from django.db.models import Count, Q
import pandas as pd
import uuid  # <--- Added this to generate unique IDs
from io import BytesIO
from datetime import date
from django.utils import timezone
from rest_framework.pagination import PageNumberPagination
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import re

# --- PAGINATION ---
class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 100

# --- UTILS ---
def get_employee_filter(user, prefix=''):
    """
    Returns a Q object to filter by the current user's employee profile.
    Matches by linked user, email, or full name.
    """
    if user.is_superuser:
        return Q()
    
    # 1. Match by direct linked user field
    q = Q(**{f"{prefix}user": user})
    
    # 2. Match by email
    if user.email:
        q |= Q(**{f"{prefix}email": user.email})
    
    # 3. Match by full name
    full_name = f"{user.first_name} {user.last_name}".strip()
    if full_name:
        q |= Q(**{f"{prefix}name__iexact": full_name})
    elif user.username:
        q |= Q(**{f"{prefix}name__iexact": user.username})
        
    return q

def _get_employee_requester(user):
    """Helper to get Employee profile for a User"""
    from .models import Employee
    # Use existing logic to find the employee profile
    q = Q(user=user)
    if user.email: q |= Q(email=user.email)
    full_name = f"{user.first_name} {user.last_name}".strip()
    if full_name: q |= Q(name__iexact=full_name)
    elif user.username: q |= Q(name__iexact=user.username)
    return Employee.objects.filter(q).first()

def clean_import_value(value, default=''):
    if pd.isna(value):
        return default
    text = str(value).strip()
    return default if text.lower() == 'nan' else text

def get_import_value(row, *column_names, default=''):
    for column_name in column_names:
        if column_name in row:
            value = clean_import_value(row.get(column_name), default='')
            if value:
                return value
    return default

def build_asset_import_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Assets'

    headers = [
        'Miczon ID',
        'Device Name',
        'Category',
        'Department',
        'Custodian',
        'Employee ID',
        'Email',
        'Specifications',
        'Remarks',
        'Status',
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 3, 14), 34)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def build_employee_import_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Employees'

    headers = [
        'Employee Name',
        'Employee ID',
        'Email',
        'Department',
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 3, 18), 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def build_next_miczon_ids(quantity):
    existing_ids = Asset.objects.values_list('miczon_id', flat=True)
    highest_number = 1000

    for miczon_id in existing_ids:
        match = re.search(r'(\d+)(?!.*\d)', str(miczon_id or ''))
        if match:
            highest_number = max(highest_number, int(match.group(1)))

    return [f"MZ-{number}" for number in range(highest_number + 1, highest_number + quantity + 1)]

# --- AUTH VIEWS ---
class CurrentUserView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)

class ScanAssetView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request, miczon_id):
        asset = Asset.objects.filter(miczon_id=miczon_id).only('id').first()
        if asset:
            return Response({"status": "found", "asset_id": asset.id})
        return Response({"status": "not_found", "miczon_id": miczon_id})

# --- VIEWSETS ---
class AssetAssignmentViewSet(viewsets.ModelViewSet):
    queryset = AssetAssignment.objects.all().order_by('-assigned_date')
    serializer_class = AssetAssignmentSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_superuser:
            # Filter assignments by the employee
            emp_filter = get_employee_filter(self.request.user, prefix='employee__')
            queryset = queryset.filter(emp_filter)
        return queryset

    def create(self, request, *args, **kwargs):
        # Strict Admin Approval Interception for Assign
        if not request.user.is_superuser:
            from .models import AssetActionRequest, Asset
            
            # Use same logic to find employee requester profile
            asset_id = request.data.get('asset')
            employee_id = request.data.get('employee')
            
            # Redirect to approval flow
            requester = _get_employee_requester(request.user)
            if not requester:
                return Response({"error": "Employee profile not found. Please contact admin to link your user account to an employee profile."}, status=400)
            
            try:
                asset = Asset.objects.get(id=asset_id)
            except Asset.DoesNotExist:
                return Response({"error": "Asset not found"}, status=400)
                
            AssetActionRequest.objects.create(
                asset=asset,
                requester=requester,
                action_type='ASSIGN',
                target_employee_id=employee_id,
                remarks=request.data.get('remarks', 'Direct Assignment Request')
            )
            return Response({"status": "Assignment request submitted for approval"}, status=201)
            
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        # Intercept Return (PATCH to assignment status)
        print(f"DEBUG: Update Assignment - User: {request.user.username}, IsSuper: {request.user.is_superuser}, Data: {request.data}")
        if not request.user.is_superuser:
            if request.data.get('status') == 'RETURNED':
                 # Redirect to approval flow for Return
                 from .models import AssetActionRequest
                 asset = self.get_object().asset
                 requester = _get_employee_requester(request.user)
                 
                 AssetActionRequest.objects.create(
                     asset=asset,
                     requester=requester,
                     action_type='RETURN',
                     remarks=f"Return Details: {request.data.get('remarks', 'N/A')}"
                 )
                 return Response({"status": "Return request submitted for approval"}, status=201)
            return Response({"error": "Only admins can modify assignments directly"}, status=403)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

class InspectionLogViewSet(viewsets.ModelViewSet):
    queryset = InspectionLog.objects.all().order_by('-date')
    serializer_class = InspectionLogSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_superuser:
            # Filter logs for assets where the user is the custodian
            emp_filter = get_employee_filter(self.request.user, prefix='asset__custodian__')
            queryset = queryset.filter(emp_filter)
        return queryset

class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.all().order_by('-created_at')
    serializer_class = AssetDetailSerializer  # Default for create/update
    pagination_class = StandardResultsSetPagination
    permission_classes = [permissions.DjangoModelPermissions]

    def get_serializer_class(self):
        """
        Use lightweight serializer for list views,
        full serializer for detail/create/update views.
        """
        if self.action == 'list':
            return AssetListSerializer
        return AssetDetailSerializer

    def get_queryset(self):
        # OPTIMIZATION: Use select_related and prefetch_related to prevent N+1 queries
        queryset = Asset.objects.all().select_related('custodian', 'department')
        
        # Prefetch assignments for list views (needed for active_assignment_id)
        if self.action in ['list', 'retrieve']:
            queryset = queryset.prefetch_related('assignments')
        
        # Only prefetch heavy relations for detail views
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related('history', 'inspectionlog_set')
        
        queryset = queryset.order_by('-created_at')
        
        # Filter parameters
        custodian = self.request.query_params.get('custodian')
        department = self.request.query_params.get('department')
        status = self.request.query_params.get('status')
        category = self.request.query_params.get('category')
        search = self.request.query_params.get('search')
        
        if custodian:
            queryset = queryset.filter(custodian_id=custodian)
        if department:
            queryset = queryset.filter(department_id=department)
        if status:
            queryset = queryset.filter(current_status=status)
        if category:
            queryset = queryset.filter(category__iexact=category)
        if search:
            # Search across multiple fields
            queryset = queryset.filter(
                Q(miczon_id__icontains=search) |
                Q(name__icontains=search) |
                Q(custodian__name__icontains=search) |
                Q(category__icontains=search)
            )
        
        # OWNER RESTRICTION: Non-superusers only see assets assigned to them
        if not self.request.user.is_superuser:
            emp_filter = get_employee_filter(self.request.user, prefix='custodian__')
            queryset = queryset.filter(emp_filter)
            
        return queryset

    def create(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            miczon_id = request.data.get('miczon_id')
            if not miczon_id:
                return Response({"error": "Miczon ID is required"}, status=400)
                
            if Asset.objects.filter(miczon_id=miczon_id).exists():
                return Response({"miczon_id": ["Asset with this Miczon ID already exists."]}, status=400)
                
            requester = _get_employee_requester(request.user)
            if not requester:
                return Response({"error": "Employee profile not found for user. Please contact admin to link your user account to an employee profile."}, status=400)
            
            AssetActionRequest.objects.create(
                requester=requester,
                action_type='ADD',
                asset_data=request.data,
                remarks='Requesting to add new asset'
            )
            return Response({"status": "Add asset request submitted for approval", "miczon_id": request.data.get('miczon_id')}, status=201)
        
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        asset = serializer.save()
        if asset.custodian and asset.current_status != 'BROKEN':
            asset.current_status = 'ASSIGNED'
            asset.save(update_fields=['current_status'])
            AssetHistory.objects.create(
                asset=asset,
                action='ADD_ASSIGN',
                to_employee=asset.custodian,
                remarks='Asset added and assigned from inventory manager.'
            )

    def perform_update(self, serializer):
        # Only superusers can directly update assets
        if not self.request.user.is_superuser:
            return # Should have been blocked by update/partial_update anyway
            
        instance = self.get_object()
        old_custodian = instance.custodian
        updated_asset = serializer.save()

        if updated_asset.custodian and updated_asset.current_status != 'BROKEN':
            updated_asset.current_status = 'ASSIGNED'
            updated_asset.save(update_fields=['current_status'])
        elif not updated_asset.custodian and updated_asset.current_status == 'ASSIGNED':
            updated_asset.current_status = 'AVAILABLE'
            updated_asset.save(update_fields=['current_status'])
        
        if updated_asset.custodian != old_custodian:
            action = 'ASSIGN' if updated_asset.custodian else 'RETURN'
            AssetHistory.objects.create(
                asset=updated_asset, action=action,
                from_employee=old_custodian, to_employee=updated_asset.custodian,
                remarks=f"Status changed to {updated_asset.current_status}. (Direct Admin Update)"
            )

    def update(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"error": "Direct updates are restricted. Please use specific action flows (Transfer/Repair/etc) which are subject to approval."}, status=403)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not request.user.is_superuser:
             return self.update(request, *args, **kwargs)
        return super().partial_update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"error": "Deletions restricted to administrators."}, status=403)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='import-template')
    def import_template(self, request):
        template = build_asset_import_template()
        return FileResponse(
            template,
            as_attachment=True,
            filename='asset_import_template.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @action(detail=False, methods=['get'], url_path='next-miczon-ids')
    def next_miczon_ids(self, request):
        try:
            quantity = int(request.query_params.get('quantity', 1))
        except (TypeError, ValueError):
            quantity = 1

        quantity = max(1, min(quantity, 500))
        return Response({"ids": build_next_miczon_ids(quantity)})

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser, FormParser])
    def import_assets(self, request):
        """
        STAGE 1: PRE-FLIGHT CHECK
        Parses Excel and reconciles data against the DB without saving.
        Returns a preview of valid rows and flagged errors.
        """
        if not request.user.is_superuser:
            return Response({"success": False, "message": "Only admins can import assets."}, status=403)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"success": False, "message": "Please upload an Excel file."}, status=400)

        try:
            df = pd.read_excel(file_obj)
        except Exception as exc:
            return Response({"success": False, "message": f"Unable to read Excel file: {exc}"}, status=400)

        if df.empty:
            return Response({"success": False, "message": "The uploaded Excel file has no asset rows."}, status=400)

        df.columns = [str(column).strip().lower() for column in df.columns]

        valid_rows = []
        errors = []
        seen_miczon_ids = set()

        for index, row in df.iterrows():
            excel_row_number = index + 2
            row_errors = []

            # 1. Reconciliation: Miczon ID (Hardware Key)
            miczon_id = get_import_value(row, 'miczon id', 'mic id', 'serial number', 'serial no', 'asset tag')
            if not miczon_id:
                row_errors.append("Missing Miczon ID")
            elif miczon_id in seen_miczon_ids:
                row_errors.append(f"Duplicate Miczon ID in file: {miczon_id}")
            seen_miczon_ids.add(miczon_id)

            # 2. Reconciliation: Device Name
            device_name = get_import_value(row, 'device name', 'device', 'name')
            if not device_name:
                row_errors.append("Missing Device Name")

            # 3. Reconciliation: Custodian (Employee Reconciliation Engine)
            # Support variations: Custodian, Custody, Assigned User, User, Employee
            custodian_name = get_import_value(row, 'custodian', 'custody', 'assigned user', 'user', 'employee name', 'employee')
            employee_id = get_import_value(row, 'employee id', 'employee code', 'emp id', 'id')

            resolved_employee = None
            if employee_id:
                resolved_employee = Employee.objects.filter(employee_id__iexact=employee_id).first()
                if not resolved_employee:
                    row_errors.append(f"Employee ID '{employee_id}' not found in system")
            elif custodian_name:
                # Fallback to name search if ID is missing, but fuzzy matches are dangerous
                resolved_employee = Employee.objects.filter(name__iexact=custodian_name).first()
                if not resolved_employee:
                    row_errors.append(f"Custodian '{custodian_name}' not found by name. (Try providing Employee ID)")

            # 4. Department logic
            dept_name = get_import_value(row, 'department')
            resolved_dept_id = None
            if resolved_employee and resolved_employee.department:
                resolved_dept_id = resolved_employee.department.id
            elif dept_name:
                dept = Department.objects.filter(name__iexact=dept_name).first()
                if dept:
                    resolved_dept_id = dept.id

            # Prepare staging data
            row_data = {
                "excel_row": excel_row_number,
                "miczon_id": miczon_id,
                "name": device_name,
                "category": get_import_value(row, 'category', 'categary'),
                "custodian_id": resolved_employee.id if resolved_employee else None,
                "custodian_name": resolved_employee.name if resolved_employee else custodian_name,
                "department_id": resolved_dept_id,
                "specifications": get_import_value(row, 'specifications', 'specs', 'details'),
                "remarks": get_import_value(row, 'remarks', 'notes'),
                "status": get_import_value(row, 'status', default='AVAILABLE').upper().replace(' ', '_')
            }

            if row_errors:
                errors.append({"row": excel_row_number, "miczon_id": miczon_id, "messages": row_errors})
            else:
                valid_rows.append(row_data)

        return Response({
            "success": True,
            "summary": {
                "total": len(df),
                "valid": len(valid_rows),
                "errors": len(errors)
            },
            "valid_rows": valid_rows,
            "errors": errors
        })

    @action(detail=False, methods=['post'], url_path='bulk-commit')
    def bulk_commit(self, request):
        """
        STAGE 2: ATOMIC EXECUTION
        Saves the validated staging data to the database.
        """
        if not request.user.is_superuser:
            return Response({"error": "Unauthorized"}, status=403)

        rows = request.data.get('rows', [])
        if not rows:
            return Response({"error": "No data to commit"}, status=400)

        created_count = 0
        updated_count = 0

        try:
            with transaction.atomic():
                for row in rows:
                    asset, created = Asset.objects.update_or_create(
                        miczon_id=row['miczon_id'],
                        defaults={
                            'name': row['name'],
                            'category': row.get('category'),
                            'department_id': row.get('department_id'),
                            'custodian_id': row.get('custodian_id'),
                            'current_status': 'ASSIGNED' if row.get('custodian_id') and row.get('status') != 'BROKEN' else row.get('status', 'AVAILABLE'),
                            'specifications': row.get('specifications', ''),
                            'remarks': row.get('remarks', ''),
                        }
                    )

                    if created:
                        created_count += 1
                        AssetHistory.objects.create(
                            asset=asset,
                            action='IMPORT',
                            remarks="Added via bulk excel import staging."
                        )
                    else:
                        updated_count += 1

            return Response({
                "success": True,
                "message": f"Bulk import complete. Created: {created_count}, Updated: {updated_count}",
                "created": created_count,
                "updated": updated_count
            })
        except Exception as e:
            return Response({"success": False, "message": f"Critical Commit Error: {str(e)}"}, status=500)
    @action(detail=True, methods=['post'])
    def transfer(self, request, pk=None):
        asset = self.get_object()
        to_employee_id = request.data.get('to_employee_id')
        remarks = request.data.get('remarks', 'Direct Transfer')
        
        try:
            to_employee = Employee.objects.get(pk=to_employee_id)
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found"}, status=400)

        # If user is not superuser, create request for approval
        if not request.user.is_superuser:
            requester = _get_employee_requester(request.user)
            if not requester:
                return Response({"error": "Employee profile not found for user"}, status=400)
            
            AssetActionRequest.objects.create(
                asset=asset,
                requester=requester,
                action_type='TRANSFER',
                target_employee=to_employee,
                remarks=remarks
            )
            return Response({"status": "Transfer request submitted for approval"})

        # Superuser: Direct transfer logic
        from_employee = asset.custodian
        active_assignment = AssetAssignment.objects.filter(asset=asset, status='ASSIGNED').first()
        if active_assignment:
            active_assignment.mark_returned(returned_by='Transfer Action', condition='Good')
        
        AssetAssignment.objects.create(
            asset=asset,
            employee=to_employee,
            remarks=remarks,
            status='ASSIGNED'
        )
        AssetHistory.objects.create(
            asset=asset,
            action="TRANSFER",
            from_employee=from_employee,
            to_employee=to_employee,
            remarks=remarks
        )
        return Response({"status": "Asset transferred successfully"})

    @action(detail=True, methods=['post'])
    def repair(self, request, pk=None):
        asset = self.get_object()
        vendor = request.data.get('vendor')
        expected_return = request.data.get('expected_return')
        issue = request.data.get('issue', 'Broken/Repair')
        
        # If user is not superuser, create request for approval
        if not request.user.is_superuser:
            requester = _get_employee_requester(request.user)
            if not requester:
                return Response({"error": "Employee profile not found for user"}, status=400)
            
            AssetActionRequest.objects.create(
                asset=asset,
                requester=requester,
                action_type='REPAIR',
                vendor=vendor,
                expected_return_date=expected_return,
                remarks=issue
            )
            return Response({"status": "Repair request submitted for approval"})

        # Superuser logic
        active_assignment = AssetAssignment.objects.filter(asset=asset, status='ASSIGNED').first()
        if active_assignment:
            active_assignment.mark_returned(returned_by='Repair Action', condition='Broken')
            
        asset.current_status = 'BROKEN'
        asset.maintenance_vendor = vendor
        asset.sent_to_repair_date = date.today()
        asset.expected_return_date = expected_return
        asset.save()
        
        AssetHistory.objects.create(
            asset=asset,
            action="REPAIR",
            remarks=f"Sent to {vendor}. Issue: {issue}"
        )
        return Response({"status": "Asset sent to repair"})

    @action(detail=True, methods=['post'])
    def return_from_repair(self, request, pk=None):
        asset = self.get_object()
        condition = request.data.get('condition', 'Good')
        remarks = request.data.get('remarks', 'Returned from repair')
        returned_by = request.data.get('returned_by', 'Admin')
        
        # 1. Update Asset Status to AVAILABLE
        asset.current_status = 'AVAILABLE'
        
        # 2. Clear repair-related fields
        asset.maintenance_vendor = ''
        asset.sent_to_repair_date = None
        asset.expected_return_date = None
        
        asset.save()
        
        # 3. Log history
        AssetHistory.objects.create(
            asset=asset,
            action="RETURN_FROM_REPAIR",
            remarks=f"Returned from repair. Condition: {condition}. {remarks}"
        )
        
        return Response({"status": "Asset returned to inventory successfully"})

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer

    def get_queryset(self):
        queryset = Employee.objects.all().annotate(
            assigned_assets_count=Count('assets', filter=Q(assets__current_status='ASSIGNED'))
        )
        department = self.request.query_params.get('department')
        if department:
            queryset = queryset.filter(department_id=department)
            
        # Security: Employees only see their own profile
        if not self.request.user.is_superuser:
            emp_filter = get_employee_filter(self.request.user)
            queryset = queryset.filter(emp_filter)
            
        return queryset

    @action(detail=False, methods=['get'], url_path='import-template')
    def import_template(self, request):
        template = build_employee_import_template()
        return FileResponse(
            template,
            as_attachment=True,
            filename='employee_import_template.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser, FormParser])
    def import_employees(self, request):
        if not request.user.is_superuser:
            return Response({"success": False, "message": "Only admins can import employees."}, status=403)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"success": False, "message": "Please upload an Excel file."}, status=400)

        try:
            df = pd.read_excel(file_obj)
        except Exception as exc:
            return Response({"success": False, "message": f"Unable to read Excel file: {exc}"}, status=400)

        if df.empty:
            return Response({"success": False, "message": "The uploaded Excel file has no employee rows."}, status=400)

        df.columns = [str(column).strip().lower() for column in df.columns]
        created_count = 0
        updated_count = 0
        skipped_rows = []

        for index, row in df.iterrows():
            excel_row_number = index + 2
            name = get_import_value(row, 'employee name', 'name', 'full name')
            emp_id = get_import_value(row, 'employee id', 'emp id', 'id', 'employee code')
            
            if not name:
                skipped_rows.append(f"Row {excel_row_number}: Missing Employee Name")
                continue
                
            if not emp_id:
                # If no ID provided, try to find by name or email, or skip
                email = get_import_value(row, 'email', 'employee email')
                existing = None
                if email:
                    existing = Employee.objects.filter(email__iexact=email).first()
                if not existing:
                    existing = Employee.objects.filter(name__iexact=name).first()
                
                if existing:
                    emp_id = existing.employee_id
                else:
                    skipped_rows.append(f"Row {excel_row_number}: Missing Employee ID")
                    continue

            department = None
            department_name = get_import_value(row, 'department')
            if department_name:
                department, _ = Department.objects.get_or_create(name=department_name)

            employee_email = get_import_value(row, 'email', 'employee email')

            _, created = Employee.objects.update_or_create(
                employee_id=emp_id,
                defaults={
                    'name': name,
                    'email': employee_email,
                    'department': department,
                },
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        imported_count = created_count + updated_count
        message = f"Successfully imported {imported_count} employee{'' if imported_count == 1 else 's'}."
        if skipped_rows:
            message += f" Skipped {len(skipped_rows)} row{'' if len(skipped_rows) == 1 else 's'}."

        return Response({
            "success": True,
            "message": message,
            "created": created_count,
            "updated": updated_count,
            "skipped": len(skipped_rows),
            "errors": skipped_rows,
        })

    @action(detail=True, methods=['get'], url_path='assigned-assets')
    def assigned_assets(self, request, pk=None):
        employee = self.get_object()
        assets = Asset.objects.filter(custodian=employee).select_related('custodian', 'department')
        serializer = AssetListSerializer(assets, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='unassign-all')
    def unassign_all(self, request, pk=None):
        if not request.user.is_superuser:
            return Response({"error": "Only admins can run employee offboarding."}, status=403)

        employee = self.get_object()
        active_assignments = AssetAssignment.objects.filter(employee=employee, status='ASSIGNED').select_related('asset')
        returned_count = 0

        for assignment in active_assignments:
            assignment.mark_returned(returned_by=f"Offboarding by {request.user.username}", condition='Good')
            AssetHistory.objects.create(
                asset=assignment.asset,
                action='OFFBOARDING_RETURN',
                from_employee=employee,
                to_employee=None,
                remarks=f"Unassigned during offboarding by {request.user.username}."
            )
            returned_count += 1

        remaining_assets = Asset.objects.filter(custodian=employee)
        for asset in remaining_assets:
            asset.custodian = None
            asset.current_status = 'AVAILABLE'
            asset.save()
            AssetHistory.objects.create(
                asset=asset,
                action='OFFBOARDING_RETURN',
                from_employee=employee,
                to_employee=None,
                remarks=f"Unassigned during offboarding by {request.user.username}."
            )
            returned_count += 1

        return Response({"status": "Employee assets unassigned", "returned_count": returned_count})

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

class AssetActionRequestViewSet(viewsets.ModelViewSet):
    queryset = AssetActionRequest.objects.all().order_by('-created_at')
    serializer_class = AssetActionRequestSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        if not self.request.user.is_superuser:
            emp_filter = get_employee_filter(self.request.user, prefix='requester__')
            queryset = queryset.filter(emp_filter)
        return queryset

    def create(self, request, *args, **kwargs):
        if not request.data.get('requester'):
            requester = _get_employee_requester(request.user)
            if not requester:
                return Response({"error": "Employee profile not found for user."}, status=400)

            data = request.data.copy()
            data['requester'] = requester.id
            data['action_type'] = data.get('action_type') or 'ASSIGN'
            data['reason_for_request'] = data.get('reason_for_request') or data.get('remarks') or ''
            data['status'] = 'PENDING'
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save(requester=requester, status='PENDING')
            return Response(serializer.data, status=201)

        if not request.user.is_superuser:
            return Response({"error": "Employees cannot create requests for another requester."}, status=403)

        return super().create(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        if not request.user.is_superuser:
            return Response({"error": "Only admins can approve requests"}, status=403)
        
        req = self.get_object()
        if req.status != 'PENDING':
            return Response({"error": "Request is already processed"}, status=400)
        
        admin_remarks = request.data.get('admin_remarks', '')
        asset = req.asset
        
        # --- PERFORM THE ACTION ---
        try:
            if req.action_type == 'ADD':
                data = req.asset_data or {}
                miczon_id = data.get('miczon_id')
                
                if not miczon_id:
                    return Response({"error": "Miczon ID is missing in request data"}, status=400)
                
                if Asset.objects.filter(miczon_id=miczon_id).exists():
                    return Response({"error": f"Asset with Miczon ID {miczon_id} already exists"}, status=400)

                department_id = data.get('department')
                department = Department.objects.filter(id=department_id).first() if department_id else None
                
                new_asset = Asset.objects.create(
                    miczon_id=miczon_id,
                    name=data.get('name', 'Unknown Asset'),
                    category=data.get('category', ''),
                    specifications=data.get('specifications', ''),
                    department=department,
                    current_status=data.get('current_status', 'AVAILABLE'),
                    remarks=data.get('remarks', f"Created via approval by {request.user.username}"),
                    # Add extra fields if they exist in data
                    maintenance_vendor=data.get('maintenance_vendor', ''),
                    sent_to_repair_date=data.get('sent_to_repair_date'),
                    expected_return_date=data.get('expected_return_date'),
                )
                req.asset = new_asset
                asset = new_asset
                
                AssetHistory.objects.create(
                    asset=asset,
                    action="ADD",
                    remarks=f"Asset added to system via approval by {request.user.username}"
                )

            elif req.action_type == 'RETURN':
                if not asset:
                    return Response({"error": "No asset associated with this request"}, status=400)
                active_assignment = AssetAssignment.objects.filter(asset=asset, status='ASSIGNED').first()
                if active_assignment:
                    active_assignment.mark_returned(returned_by=f"Approved by {request.user.username}", condition='Good')
                else:
                    # Fallback if no active assignment found but request exists
                    asset.custodian = None
                    asset.current_status = 'AVAILABLE'
                    asset.save()
                
                AssetHistory.objects.create(
                    asset=asset,
                    action="RETURN",
                    remarks=f"Return request approved by {request.user.username}"
                )
        
            elif req.action_type == 'ASSIGN' or req.action_type == 'TRANSFER':
                if not asset:
                    if req.requested_device_type:
                        # Device-only requests are approvals to provision hardware; assignment happens once an asset is selected.
                        pass
                    else:
                        return Response({"error": "No asset associated with this request"}, status=400)
                elif not req.target_employee:
                    return Response({"error": "Target employee is missing in request"}, status=400)
                else:
                    from_employee = asset.custodian
                    to_employee = req.target_employee
                    
                    # Return current assignment if transfer or if already assigned
                    active_assignment = AssetAssignment.objects.filter(asset=asset, status='ASSIGNED').first()
                    if active_assignment:
                        active_assignment.mark_returned(returned_by=f"System (Prior to {req.action_type} Approval)", condition='Good')
                    
                    AssetAssignment.objects.create(
                        asset=asset,
                        employee=to_employee,
                        remarks=req.remarks,
                        status='ASSIGNED'
                    )
                    AssetHistory.objects.create(
                        asset=asset,
                        action=req.action_type,
                        from_employee=from_employee,
                        to_employee=to_employee,
                        remarks=req.remarks or f"{req.action_type} Approved by {request.user.username}"
                    )
            
            elif req.action_type == 'REPAIR':
                if not asset:
                    return Response({"error": "No asset associated with this request"}, status=400)
                
                active_assignment = AssetAssignment.objects.filter(asset=asset, status='ASSIGNED').first()
                if active_assignment:
                    active_assignment.mark_returned(returned_by=f"Repair Approved by {request.user.username}", condition='Broken')
                    
                asset.current_status = 'BROKEN'
                asset.maintenance_vendor = req.vendor
                asset.sent_to_repair_date = date.today()
                asset.expected_return_date = req.expected_return_date
                asset.save()
                
                AssetHistory.objects.create(
                    asset=asset,
                    action="REPAIR",
                    remarks=f"Repair Approved. Vendor: {req.vendor}. Info: {req.remarks}"
                )
        except Exception as e:
            return Response({"error": f"Failed to process approval: {str(e)}"}, status=500)

        # Update Request Status
        req.status = 'APPROVED'
        req.admin_remarks = admin_remarks
        req.processed_at = timezone.now()
        req.processed_by = request.user
        req.save()
        
        return Response({"status": "Request approved and action performed"})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        if not request.user.is_superuser:
            return Response({"error": "Only admins can reject requests"}, status=403)
        
        req = self.get_object()
        if req.status != 'PENDING':
            return Response({"error": "Request is already processed"}, status=400)
            
        admin_remarks = request.data.get('admin_remarks', '')
        req.status = 'REJECTED'
        req.admin_remarks = admin_remarks
        req.processed_at = timezone.now()
        req.processed_by = request.user
        req.save()
        
        return Response({"status": "Request rejected"})

class HealthCheckSessionViewSet(viewsets.ModelViewSet):
    queryset = HealthCheckSession.objects.all().order_by('-created_at')
    serializer_class = HealthCheckSessionSerializer

    def get_queryset(self):
        queryset = super().get_queryset().annotate(response_count=Count('responses'))
        if self.request.user.is_superuser:
            return queryset

        employee = _get_employee_requester(self.request.user)
        if not employee:
            return queryset.none()

        submitted_sessions = HealthCheckResponse.objects.filter(employee=employee).values('session_id')
        return queryset.filter(Q(status='OPEN') | Q(id__in=submitted_sessions)).distinct()

    @action(detail=False, methods=['post'], url_path='trigger-global')
    def trigger_global(self, request):
        if not request.user.is_superuser:
            return Response({"error": "Only admins can start monthly inspections."}, status=403)

        title = request.data.get('title') or f"Monthly Hardware Inspection {timezone.now().date()}"
        session = HealthCheckSession.objects.create(title=title, triggered_by=request.user)
        serializer = self.get_serializer(session)
        assigned_assets = Asset.objects.filter(current_status='ASSIGNED', custodian__isnull=False).count()
        return Response({
            "status": "Monthly inspection started",
            "assigned_assets": assigned_assets,
            "session": serializer.data,
        }, status=201)

    @action(detail=True, methods=['post'], url_path='close')
    def close(self, request, pk=None):
        if not request.user.is_superuser:
            return Response({"error": "Only admins can close health checks."}, status=403)

        session = self.get_object()
        session.status = 'CLOSED'
        session.closed_at = timezone.now()
        session.save()
        return Response({"status": "Health check closed"})

    @action(detail=True, methods=['get'], url_path='responses')
    def responses(self, request, pk=None):
        session = self.get_object()
        responses = HealthCheckResponse.objects.filter(session=session).select_related('employee', 'asset')
        if not request.user.is_superuser:
            employee = _get_employee_requester(request.user)
            responses = responses.filter(employee=employee)
        serializer = HealthCheckResponseSerializer(responses, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='pending-assets')
    def pending_assets(self, request, pk=None):
        session = self.get_object()
        employee = _get_employee_requester(request.user)

        if request.user.is_superuser and request.query_params.get('employee'):
            employee = Employee.objects.filter(id=request.query_params.get('employee')).first()

        if not employee:
            return Response([])

        responded_assets = HealthCheckResponse.objects.filter(
            session=session,
            employee=employee
        ).values_list('asset_id', flat=True)
        assets = Asset.objects.filter(custodian=employee).exclude(id__in=responded_assets)
        serializer = AssetListSerializer(assets, many=True, context={'request': request})
        return Response(serializer.data)

class HealthCheckResponseViewSet(viewsets.ModelViewSet):
    queryset = HealthCheckResponse.objects.all().select_related('session', 'employee', 'asset').order_by('-submitted_at')
    serializer_class = HealthCheckResponseSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        asset = self.request.query_params.get('asset')
        session = self.request.query_params.get('session')
        if asset:
            queryset = queryset.filter(asset_id=asset)
        if session:
            queryset = queryset.filter(session_id=session)

        if not self.request.user.is_superuser:
            employee = _get_employee_requester(self.request.user)
            queryset = queryset.filter(employee=employee)

        return queryset

    def create(self, request, *args, **kwargs):
        employee = _get_employee_requester(request.user)
        if not employee:
            return Response({"error": "Employee profile not found for user."}, status=400)

        data = request.data.copy()
        data['employee'] = employee.id
        asset_id = data.get('asset')
        session_id = data.get('session')

        if not Asset.objects.filter(id=asset_id, custodian=employee).exists() and not request.user.is_superuser:
            return Response({"error": "You can only submit checks for hardware assigned to you."}, status=403)

        if not HealthCheckSession.objects.filter(id=session_id, status='OPEN').exists():
            return Response({"error": "This health check session is not open."}, status=400)

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save(employee=employee)
        return Response(serializer.data, status=201)

    @action(detail=False, methods=['post'], url_path='bulk-submit')
    def bulk_submit(self, request):
        employee = _get_employee_requester(request.user)
        if not employee:
            return Response({"error": "Employee profile not found for user."}, status=400)

        session_id = request.data.get('session')
        responses = request.data.get('responses') or []
        if not responses:
            return Response({"error": "No health check responses were provided."}, status=400)

        session = HealthCheckSession.objects.filter(id=session_id, status='OPEN').first()
        if not session:
            return Response({"error": "This health check session is not open."}, status=400)

        asset_ids = [response.get('asset') for response in responses if response.get('asset')]
        if len(asset_ids) != len(responses):
            return Response({"error": "Every health check response must include an asset."}, status=400)

        assigned_asset_ids = set(Asset.objects.filter(id__in=asset_ids, custodian=employee).values_list('id', flat=True))
        if not request.user.is_superuser and assigned_asset_ids != set(asset_ids):
            return Response({"error": "You can only submit checks for hardware assigned to you."}, status=403)

        saved_responses = []
        with transaction.atomic():
            for response in responses:
                data = {
                    'session': session.id,
                    'employee': employee.id,
                    'asset': response.get('asset'),
                    'screen_condition': response.get('screen_condition') or 'GOOD',
                    'battery_life': response.get('battery_life') or 'GOOD',
                    'physical_condition': response.get('physical_condition') or 'GOOD_MINOR_WEAR',
                    'power_boot_status': response.get('power_boot_status') or 'BOOTS_NORMALLY',
                    'ports_connectors': response.get('ports_connectors') or 'ALL_FUNCTIONAL',
                    'network_functionality': response.get('network_functionality') or 'CONNECTS_NORMALLY',
                    'asset_tag_status': response.get('asset_tag_status') or 'INTACT_SCANNABLE',
                    'performance_rating': response.get('performance_rating') or 4,
                    'comments': response.get('comments') or '',
                }
                existing_response = HealthCheckResponse.objects.filter(
                    session=session,
                    employee=employee,
                    asset_id=response.get('asset')
                ).first()
                serializer = self.get_serializer(existing_response, data=data)
                serializer.is_valid(raise_exception=True)
                health_response = serializer.save(employee=employee)
                saved_responses.append(health_response)

        serializer = self.get_serializer(saved_responses, many=True)
        return Response(serializer.data, status=201)

class ReportsViewSet(viewsets.ViewSet):
    """
    Reporting and Analytics ViewSet
    """

    @action(detail=False, methods=['get'], url_path='dashboard-stats')
    def dashboard_stats(self, request):
        queryset = Asset.objects.all()
        if not request.user.is_superuser:
            emp_filter = get_employee_filter(request.user, prefix='custodian__')
            queryset = queryset.filter(emp_filter)

        stats = queryset.aggregate(
            total_assets=Count('id'),
            total_assigned=Count('id', filter=Q(current_status='ASSIGNED')),
            total_unassigned=Count('id', filter=Q(current_status='AVAILABLE')),
            total_repair=Count('id', filter=Q(current_status='BROKEN') | Q(current_status='IN_REPAIR'))
        )
        return Response(stats)

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        assets = Asset.objects.all()
        if not request.user.is_superuser:
            emp_filter = get_employee_filter(request.user, prefix='custodian__')
            assets = assets.filter(emp_filter)

        category_rows = assets.values('category').annotate(count=Count('id')).order_by('category')
        active_requests = AssetActionRequest.objects.filter(status='PENDING')
        pending_health_checks = HealthCheckSession.objects.filter(status='OPEN')

        if not request.user.is_superuser:
            employee = _get_employee_requester(request.user)
            active_requests = active_requests.filter(requester=employee)
            if employee:
                assigned_assets = Asset.objects.filter(custodian=employee)
                submitted = HealthCheckResponse.objects.filter(employee=employee, asset__in=assigned_assets).values('session_id')
                pending_health_checks = pending_health_checks.exclude(id__in=submitted)
            else:
                pending_health_checks = pending_health_checks.none()

        return Response({
            "total_devices": assets.count(),
            "laptops": assets.filter(category__icontains='laptop').count(),
            "mobiles": assets.filter(Q(category__icontains='mobile') | Q(category__icontains='phone')).count(),
            "accessories": assets.filter(category__icontains='accessor').count(),
            "assigned": assets.filter(current_status='ASSIGNED').count(),
            "available": assets.filter(current_status='AVAILABLE').count(),
            "repair": assets.filter(Q(current_status='BROKEN') | Q(current_status='IN_REPAIR')).count(),
            "active_requests": active_requests.count(),
            "pending_health_checks": pending_health_checks.count(),
            "category_breakdown": list(category_rows),
        })

    @action(detail=False, methods=['get'], url_path='category-breakdown')
    def category_breakdown(self, request):
        queryset = Asset.objects.all()
        if not request.user.is_superuser:
            emp_filter = get_employee_filter(request.user, prefix='custodian__')
            queryset = queryset.filter(emp_filter)

        # Return asset counts grouped by Category.
        data = queryset.values('category').annotate(
            count=Count('id'),
            assigned=Count('id', filter=Q(current_status='ASSIGNED')),
            available=Count('id', filter=Q(current_status='AVAILABLE'))
        ).order_by('category')
        
        return Response(data)

    @action(detail=False, methods=['get'], url_path='health-compliance')
    def health_compliance(self, request):
        sessions = HealthCheckSession.objects.all().annotate(response_count=Count('responses')).order_by('-created_at')
        session_id = request.query_params.get('session')
        session = sessions.filter(id=session_id).first() if session_id else sessions.filter(status='OPEN').first()
        if not session:
            session = sessions.first()

        if not session:
            return Response({
                "session": None,
                "summary": {
                    "target_assets": 0,
                    "completed_assets": 0,
                    "pending_assets": 0,
                    "pending_employees": 0,
                    "completion_rate": 0,
                    "critical_alerts": 0,
                },
                "pending_by_employee": [],
                "department_summary": [],
                "responses": [],
            })

        target_assets = Asset.objects.filter(current_status='ASSIGNED', custodian__isnull=False).select_related('custodian', 'department', 'custodian__department')
        if not request.user.is_superuser:
            employee = _get_employee_requester(request.user)
            target_assets = target_assets.filter(custodian=employee) if employee else target_assets.none()
        target_assets = list(target_assets)
        target_asset_ids = [asset.id for asset in target_assets]

        response_qs = HealthCheckResponse.objects.filter(session=session, asset_id__in=target_asset_ids).select_related('employee', 'employee__department', 'asset', 'asset__department')
        if not request.user.is_superuser:
            employee = _get_employee_requester(request.user)
            response_qs = response_qs.filter(employee=employee) if employee else response_qs.none()

        responded_asset_ids = set(response_qs.values_list('asset_id', flat=True))
        pending_assets = [asset for asset in target_assets if asset.id not in responded_asset_ids]

        pending_by_employee = {}
        for asset in pending_assets:
            employee = asset.custodian
            if not employee:
                continue
            row = pending_by_employee.setdefault(employee.id, {
                "employee_id": employee.id,
                "employee_name": employee.name,
                "employee_code": employee.employee_id,
                "email": employee.email,
                "department": employee.department.name if employee.department else (asset.department.name if asset.department else "Unassigned"),
                "pending_count": 0,
                "assets": [],
            })
            row["pending_count"] += 1
            row["assets"].append({
                "id": asset.id,
                "name": asset.name,
                "miczon_id": asset.miczon_id,
                "category": asset.category,
            })

        department_summary = {}
        for asset in target_assets:
            department_name = asset.custodian.department.name if asset.custodian and asset.custodian.department else (asset.department.name if asset.department else "Unassigned")
            row = department_summary.setdefault(department_name, {"department": department_name, "target": 0, "completed": 0, "pending": 0})
            row["target"] += 1
            if asset.id in responded_asset_ids:
                row["completed"] += 1
            else:
                row["pending"] += 1

        total_targets = len(target_assets)
        completed_assets = len(responded_asset_ids)
        completion_rate = round((completed_assets / total_targets) * 100) if total_targets else 0

        return Response({
            "session": HealthCheckSessionSerializer(session).data,
            "summary": {
                "target_assets": total_targets,
                "completed_assets": completed_assets,
                "pending_assets": max(total_targets - completed_assets, 0),
                "pending_employees": len(pending_by_employee),
                "completion_rate": completion_rate,
                "critical_alerts": response_qs.filter(performance_rating__lt=3).count(),
            },
            "pending_by_employee": sorted(pending_by_employee.values(), key=lambda row: (-row["pending_count"], row["employee_name"])),
            "department_summary": sorted(department_summary.values(), key=lambda row: row["department"]),
            "responses": HealthCheckResponseSerializer(response_qs.order_by('-submitted_at')[:100], many=True).data,
        })

    @action(detail=False, methods=['get'], url_path='custom-export')
    def custom_export(self, request):
        # OPTIMIZATION: Eager load relationships
        queryset = Asset.objects.all().select_related('custodian', 'department').prefetch_related(
            'history', 'assignments', 'inspectionlog_set'
        )
        
        # Filtering parameters
        dept_id = request.query_params.get('department')
        location = request.query_params.get('location') # Assuming this maps to Floor or similar
        category = request.query_params.get('category')
        status = request.query_params.get('status')

        if dept_id:
            queryset = queryset.filter(department_id=dept_id)
        
        if location:
            # Assuming location filters by Department Floor as there is no specific Location model
            queryset = queryset.filter(department__floor__icontains=location)
        
        if category:
            queryset = queryset.filter(category__icontains=category)

        if status:
            queryset = queryset.filter(current_status=status)

        # Owner restricted export
        if not request.user.is_superuser:
            emp_filter = get_employee_filter(request.user, prefix='custodian__')
            queryset = queryset.filter(emp_filter)

        serializer = AssetSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)

# --- EXCEL UPLOAD LOGIC ---
class UploadAssetsView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES['file']
        try:
            # 1. Read File
            if file_obj.name.endswith('.csv'):
                df = pd.read_csv(file_obj)
            else:
                df = pd.read_excel(file_obj)

            df.columns = [str(c).strip().lower() for c in df.columns]
            print("LOWERCASE COLUMNS:", df.columns.tolist())

            count = 0
            for index, row in df.iterrows():
                # A. Get Miczon ID (Asset Tag)
                m_id = str(row.get('miczon id', row.get('mic id', row.get('serial no', '')))).strip()
                if not m_id or m_id.lower() == 'nan':
                    continue 

                # B. Handle Department
                dept_name = str(row.get('department', 'General')).strip()
                department, _ = Department.objects.get_or_create(name=dept_name)

                # C. Handle Custodian (SAFE METHOD)
                cust_name = str(row.get('custodian', row.get('user', ''))).strip()
                custodian = None
                
                if cust_name and cust_name.lower() != 'nan':
                    # 1. Try to find existing employee by name
                    custodian = Employee.objects.filter(name__iexact=cust_name).first()
                    
                    # 2. If not found, create a new one with a RANDOM UNIQUE ID
                    if not custodian:
                        unique_emp_id = f"EMP-{uuid.uuid4().hex[:6].upper()}" # e.g., EMP-9A4F12
                        custodian = Employee.objects.create(
                            name=cust_name,
                            employee_id=unique_emp_id,
                            department=department
                        )

                # D. Save Asset
                Asset.objects.update_or_create(
                    miczon_id=m_id,
                    defaults={
                        'name': row.get('device name', row.get('device', 'Unknown Device')),
                        'category': row.get('category', row.get('categary', '')), 
                        'specifications': row.get('specifications', row.get('specs', row.get('details', row.get('hardware info', row.get('technical specs', ''))))),
                        'remarks': row.get('remarks / notes', row.get('remarks', row.get('notes', ''))),
                        'department': department,
                        'custodian': custodian,
                        'current_status': 'ASSIGNED' if custodian else 'AVAILABLE'
                    }
                )
                count += 1
            
            return Response({"status": "success", "message": f"Successfully imported {count} assets!"})

        except Exception as e:
            print("--------------------------------------------------")
            print("CRITICAL UPLOAD ERROR:", e)
            print("--------------------------------------------------")
            return Response({"status": "error", "message": f"Error: {str(e)}"}, status=400)
